#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2021/7/12 14:55
# @Author  : hzy
# @File    : excel_split.py
# @Software: PyCharm

import pandas as pd
from openpyxl import load_workbook
import numpy as np
from os import path
from copy import copy
import os


class ExcelSplit(object):

    def __init__(self, filename: str):
        self.filename = filename
        self.xlsx = pd.ExcelFile(self.filename)
        self.wb = load_workbook(self.filename)
        self._writer = None

    @staticmethod
    def get_bounds(bounds):
        start_row, start_col, end_row, end_col = bounds
        return start_row - 1, start_col - 2, end_row, end_col - 1

    def get_notna_value(self, data):
        if isinstance(data, pd.DataFrame):
            data = data[data.count(axis=1) >= 1]
            return np.nan if data.empty else self.get_notna_value(data[data.count(axis=1) >= 1].iloc[0])
        if isinstance(data, pd.Series):
            return data[data.notna()].values[0]

    def unmerge_cells(self, df, info):
        for x in info.merged_cells.ranges:
            start_row, start_col, end_row, end_col = self.get_bounds(x.bounds)
            if start_col < 0:
                continue
            # print(x.bounds)
            df.iloc[start_col:end_col, start_row:end_row] = self.get_notna_value(
                df.iloc[start_col:end_col, start_row:end_row])
        return df

    def get_new_name(self):
        filename, suffix = path.splitext(self.filename)
        return f'{filename}_new{suffix}'

    @property
    def writer(self):
        if not self._writer:
            options = dict()
            options['strings_to_formulas'] = False
            options['strings_to_urls'] = False
            options['strings_to_numbers'] = False
            new_filename = self.get_new_name()
            writer = pd.ExcelWriter(new_filename, engine='openpyxl', options=options)
            book = load_workbook(self.filename)
            writer.book = book
            self._writer = writer
        return self._writer

    def hidden_sheet(self, new_sheet, name):
        self.writer.sheets[new_sheet].sheet_state = self.wb[name].sheet_state

    def format_columns(self, new_sheet, name):
        sheet = self.writer.sheets[new_sheet]
        origin_sheet = self.wb[name]
        for col in self.wb[name].column_dimensions:
            sheet.column_dimensions[col] = copy(origin_sheet.column_dimensions[col])
            # sheet.column_dimensions[col].parent = sheet

    def format_rows(self, new_sheet, name):
        sheet = self.writer.sheets[new_sheet]
        origin_sheet = self.wb[name]
        for row in self.wb[name].row_dimensions:
            sheet.row_dimensions[row] = copy(origin_sheet.row_dimensions[row])
            # sheet.row_dimensions[row].parent = sheet

    def run(self):
        result = {}
        names = self.xlsx.sheet_names[:]
        for name in names:
            df = pd.read_excel(self.xlsx, sheet_name=name, engine='openpyxl')
            sheet_info = self.wb[name]
            df = self.unmerge_cells(df, sheet_info)
            self.fill_columns_label(df, name)
            result[name] = df.copy()

        for name in result:
            try:
                print(name)
                new_sheet = f'{name}_1'
                if len(name) >= 31:
                    new_sheet = name[:29] + '_1'
                result[name].to_excel(self.writer, new_sheet, index=False)
                self.hidden_sheet(new_sheet, name)
                self.format_rows(new_sheet, name)
                self.format_columns(new_sheet, name)

            except Exception as e:
                print(e)
        # self.format_book()
        self.writer.close()
        self.wb.close()
        self.xlsx.close()

    def format_book(self):
        self.writer.book._alignments = self.wb._alignments
        self.writer.book._named_style = self.wb._named_styles
        self.writer.book._fonts = self.wb._fonts

    @staticmethod
    def fill_columns_label(data, name):
        columns = list(data.columns)
        pre_name = name
        for i in range(len(columns)):
            if not isinstance(columns[i], str):
                continue
            if columns[i][:8] == 'Unnamed:':
                columns[i] = pre_name
            pre_name = columns[i]
        data.columns = columns


def run_excel(p):
    paths = os.listdir(p)
    for _path in paths:
        _path = f'{p}/{_path}'
        if os.path.isfile(_path):
            if _path.split('.')[-1] == 'xlsx':
                print(_path)
                es = ExcelSplit(_path)
                es.run()
        else:
            run_excel(_path)


if __name__ == '__main__':
    for i in range(1, 5):
        es = ExcelSplit(f'test{i}.xlsx')
        es.run()

    # run_excel('/Users/huzhenyu/Downloads/20210501')
    # run_excel('/Users/huzhenyu/Downloads/4月份')
